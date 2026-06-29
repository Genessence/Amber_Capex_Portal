#!/usr/bin/env python3
"""Parse Capex FY 2026-27 RAC Plants workbook and emit brownfield seed TS."""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Capex FY 2026-27 RAC Plants final 18 April.xlsx"
OUT = ROOT / "src/lib/brownFieldSeedData.ts"

SHEET_PLANT = {
    "DDN-4": "ddn_4",
    "DDN-5": "ddn_5",
    "DDN-6": "ddn_6",
    "JJR P1": "jhajjar_p1",
    "JJR P2": "jhajjar_p2",
    "SUPA": "supa",
    "Rudrapur": "rudrapur",
    "Sricity-1": "sircity_1",
    "Sricity-2": "sircity_2",
}

FY = "2026-27"
CR = 10_000_000


def norm_head(h: str | None) -> str:
    if not h:
        return ""
    s = str(h).strip()
    s = re.sub(r", If any$", "", s, flags=re.I)
    mapping = {
        "genral": "General",
        "general": "General",
        "new business": "New Business",
        "misc.": "Misc.",
        "misc": "Misc.",
    }
    low = s.lower()
    for k, v in mapping.items():
        if low == k or low.startswith(k):
            if k == "new business":
                return "New Business"
            return v
    return s


def is_total_row(sno, sub) -> bool:
    s = str(sno).strip().lower() if sno is not None else ""
    sub_s = str(sub).strip().lower() if sub is not None else ""
    if s in ("total",) or s.startswith("total "):
        return True
    if sub_s in ("total", "total value required"):
        return True
    return False


def normalize_rate_rs(rate_val, qty_val, total_cr: float | None) -> int | None:
    if rate_val is None or rate_val == "":
        return None
    try:
        rate = float(rate_val)
    except (TypeError, ValueError):
        return None
    if not (rate == rate):  # NaN
        return None

    qty = None
    if qty_val is not None and qty_val != "":
        try:
            qty = float(qty_val)
        except (TypeError, ValueError):
            qty = None

    if rate > 0 and rate < 100:
        if qty and qty > 0 and total_cr is not None:
            expected = rate * qty
            if abs(expected - total_cr) < 0.02:
                return round((total_cr / qty) * CR)
        if (qty is None or qty == 1) and total_cr is not None and abs(rate - total_cr) < 0.001:
            return round(total_cr * CR)
        return round(rate * CR)
    return round(rate)


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if any(v.lower().startswith("s.no") for v in vals):
            return i, list(row)
    return None, None


def col_index(headers, *needles):
    lower = [str(h).lower().replace("\n", " ") if h is not None else "" for h in headers]
    for idx, h in enumerate(lower):
        if all(n in h for n in needles):
            return idx
    return None


def parse_sheet(ws, plant: str) -> list[dict]:
    header_row, headers = find_header_row(ws)
    if not header_row:
        return []

    lower = [str(h).lower().replace("\n", " ") if h is not None else "" for h in headers]
    c_sno = col_index(headers, "s.no")
    c_head = col_index(headers, "head")
    c_sub = col_index(headers, "sub")
    c_dept = col_index(headers, "department") or col_index(headers, "departments")
    c_rate = col_index(headers, "rate")
    c_qty = next((i for i, h in enumerate(lower) if "qty" in h), None)
    tc_cols = [i for i, h in enumerate(lower) if "total cost" in h and "cr" in h]
    c_total = tc_cols[-1] if len(tc_cols) > 1 else (tc_cols[0] if tc_cols else None)
    c_reason = col_index(headers, "reason")
    c_benefits = col_index(headers, "benefit")
    c_roi = col_index(headers, "roi")

    items = []
    current_head = ""

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(row)

        def get(c):
            return row[c] if c is not None and c < len(row) else None

        sno_raw = get(c_sno)
        sub_raw = get(c_sub)
        head_raw = get(c_head)

        if sno_raw is None and (sub_raw is None or sub_raw == ""):
            continue

        if head_raw not in (None, ""):
            current_head = norm_head(head_raw)

        sub = str(sub_raw).strip() if sub_raw is not None else ""
        sno_str = str(sno_raw).strip() if sno_raw is not None else ""

        total_cr_val = get(c_total)
        total_cost = float(total_cr_val) if total_cr_val is not None and str(total_cr_val) not in ("", "#DIV/0!") else 0.0

        is_uc = plant == "jhajjar_p1" and sub and re.search(r"urban company|^uc$", sub, re.I)

        is_sricity_placeholder = (
            plant == "sircity_2"
            and current_head == "New Business"
            and not sub
            and total_cost > 1
        )
        # Row 39 is a sheet-level total rollup, not an additional budget line.
        if is_sricity_placeholder:
            continue

        if is_total_row(sno_raw, sub_raw) and not is_uc:
            continue
        if not sub and not is_uc:
            continue

        qty_raw = get(c_qty)
        qty = None
        if qty_raw is not None and str(qty_raw).strip() not in ("", "-", "#DIV/0!"):
            try:
                qty = float(qty_raw)
            except (TypeError, ValueError):
                qty = None
        rate_rs = normalize_rate_rs(get(c_rate), qty_raw, total_cost)
        rate_cr = (rate_rs / CR) if rate_rs is not None else 0.0

        reason = get(c_reason)
        benefits = get(c_benefits)
        roi_raw = get(c_roi)
        roi = None
        if roi_raw is not None and str(roi_raw).strip() not in ("", "-", "#DIV/0!"):
            roi = str(roi_raw).strip()

        dept_raw = get(c_dept)
        department = str(dept_raw).strip() if dept_raw is not None else ""

        item = {
            "plant": plant,
            "head": current_head or "General",
            "department": department,
            "division": "Other Brown Field",
            "subParticulars": sub
            if not is_uc
            else "UC (Urban Company — already approved Nov 2025, balance shifted to FY26-27)",
            "rate": rate_cr,
            "totalCost": round(total_cost, 10),
            "fieldType": "brown_field",
            "fy": FY,
        }
        if sno_str and not is_uc:
            item["sNo"] = sno_str
        if is_uc:
            item["sNo"] = "NB1"
        if rate_rs is not None:
            item["rateRs"] = rate_rs
        if qty is not None:
            item["qty"] = qty
        if reason:
            item["reasonForRequirement"] = str(reason).strip()
        if benefits:
            item["benefits"] = str(benefits).strip()
        if roi:
            item["roi"] = roi

        items.append(item)

    return items


def esc_ts(s: str) -> str:
    return json.dumps(s, ensure_ascii=False)


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    all_items: list[dict] = []

    for sheet_name, plant in SHEET_PLANT.items():
        ws = wb[sheet_name]
        rows = parse_sheet(ws, plant)
        all_items.extend(rows)
        total = sum(r["totalCost"] for r in rows)
        print(f"{sheet_name} ({plant}): {len(rows)} rows, {total:.3f} Cr")

    totals: dict[str, float] = {}
    for item in all_items:
        totals[item["plant"]] = totals.get(item["plant"], 0) + item["totalCost"]
    print("\nPlant totals:", {k: round(v, 3) for k, v in totals.items()})
    print("Grand total:", round(sum(totals.values()), 3))

    lines = [
        "import type { CapexMasterItem } from './types';",
        "",
        "/** Bump when Brown Field workbook seed is regenerated — triggers one-time localStorage migration. */",
        "export const BROWNFIELD_SEED_VERSION = 'fy2026_27_rac_plants';",
        "",
        "/** FY 2026-27 Brown Field RAC plant master — generated from workbook. Do not edit by hand. */",
        "export const brownFieldSeedData: CapexMasterItem[] = [",
    ]

    for idx, item in enumerate(all_items):
        item_id = f"cm-bf-{item['plant']}-{idx + 1:04d}"
        parts = [
            f"id: {esc_ts(item_id)}",
            "fieldType: 'brown_field'",
            f"fy: {esc_ts(FY)}",
            f"plant: {esc_ts(item['plant'])}",
            f"division: {esc_ts(item.get('division', 'Other Brown Field'))}",
            f"head: {esc_ts(item['head'])}",
            f"department: {esc_ts(item.get('department', ''))}",
            f"subParticulars: {esc_ts(item['subParticulars'])}",
            f"rate: {item['rate']}",
            f"totalCost: {item['totalCost']}",
        ]
        if item.get("sNo"):
            parts.append(f"sNo: {esc_ts(item['sNo'])}")
        if item.get("rateRs") is not None:
            parts.append(f"rateRs: {item['rateRs']}")
        if item.get("qty") is not None:
            parts.append(f"qty: {item['qty']}")
        if item.get("reasonForRequirement"):
            parts.append(f"reasonForRequirement: {esc_ts(item['reasonForRequirement'])}")
        if item.get("benefits"):
            parts.append(f"benefits: {esc_ts(item['benefits'])}")
        if item.get("roi"):
            parts.append(f"roi: {esc_ts(item['roi'])}")

        lines.append(f"  {{ {', '.join(parts)} }},")

    lines.append("];")
    lines.append("")
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nWrote {len(all_items)} items to {OUT}")


if __name__ == "__main__":
    main()
